"""
数据图表生成器 v4.6.0
功能：分析 Excel 数据特征 → 推荐图表类型 → 生成图表嵌入 Excel

v4.6.0 变更:
  - 🎯 数据特征自动分析（时间序列/类别对比/占比/相关性/多变量）
  - 🎯 智能图表推荐（折线图/柱状图/饼图/散点图/热力图/面积图）
  - 🎯 图表自动生成并嵌入 Excel（openpyxl 原生图表）
  - 🎯 硬件自适应（低配禁用复杂图表渲染）
  - 🎯 批量处理模式（批量分析多 Sheet 并推荐）
"""

import os
import sys
import json
import math
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent))

try:
    import openpyxl
    from openpyxl.chart import (
        LineChart, BarChart, PieChart, ScatterChart, AreaChart,
        Reference, Series
    )
    from openpyxl.chart.label import DataLabelList
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from wps_common import safe_path, get_hardware_info, with_retry
except ImportError:
    def safe_path(p): return Path(p)
    def get_hardware_info(): return {"cpu_cores": 4, "memory_gb": 8, "level": "medium"}
    def with_retry(f): return f


class DataAnalyzer:
    """数据特征分析器"""
    
    def __init__(self, filepath: str, sheet: str = "Sheet1"):
        self.filepath = filepath
        self.sheet = sheet
        self.wb = None
        self.ws = None
        self.headers = []
        self.data_rows = 0
        self.col_types = {}  # col_index -> type
        self.col_stats = {}  # col_index -> stats
    
    def load(self) -> bool:
        """加载 Excel 数据"""
        try:
            self.wb = openpyxl.load_workbook(self.filepath, read_only=True, data_only=True)
            if self.sheet not in self.wb.sheetnames:
                self.sheet = self.wb.sheetnames[0]
            self.ws = self.wb[self.sheet]
            
            # 读取表头
            first_row = next(self.ws.iter_rows(min_row=1, max_row=1, values_only=True), [])
            self.headers = [str(h) if h else f"列{i+1}" for i, h in enumerate(first_row)]
            
            # 统计行数和列类型
            self.data_rows = 0
            col_samples = {i: [] for i in range(len(self.headers))}
            
            for row in self.ws.iter_rows(min_row=2, values_only=True):
                self.data_rows += 1
                for i, val in enumerate(row):
                    if i < len(self.headers) and val is not None:
                        col_samples[i].append(val)
                        if len(col_samples[i]) >= 20:  # 最多采样 20 个
                            continue
            
            # 判断列类型
            for i in range(len(self.headers)):
                samples = col_samples[i]
                self.col_types[i] = self._detect_type(samples)
                self.col_stats[i] = self._calc_stats(samples, self.col_types[i])
            
            return True
        except Exception as e:
            return False
    
    def _detect_type(self, samples: List) -> str:
        """检测列数据类型"""
        if not samples:
            return "empty"
        
        date_count = 0
        num_count = 0
        str_count = 0
        
        for val in samples:
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                num_count += 1
            elif isinstance(val, datetime):
                date_count += 1
            elif isinstance(val, str):
                try:
                    float(val.replace(",", ""))
                    num_count += 1
                except (ValueError, TypeError):
                    # 尝试日期解析
                    try:
                        datetime.strptime(val[:10], "%Y-%m-%d")
                        date_count += 1
                    except (ValueError, TypeError):
                        str_count += 1
        
        total = len(samples)
        if date_count / total > 0.6:
            return "datetime"
        elif num_count / total > 0.7:
            return "numeric"
        else:
            return "category"
    
    def _calc_stats(self, samples: List, col_type: str) -> Dict:
        """计算列统计信息"""
        stats = {"count": len(samples), "unique": 0, "min": None, "max": None}
        
        if col_type == "numeric":
            nums = [float(v) for v in samples if isinstance(v, (int, float))]
            if nums:
                stats["min"] = min(nums)
                stats["max"] = max(nums)
                stats["unique"] = len(set(nums))
        elif col_type in ("category", "datetime"):
            stats["unique"] = len(set(str(v) for v in samples))
        
        return stats
    
    def get_profile(self) -> Dict:
        """获取数据画像"""
        return {
            "file": self.filepath,
            "sheet": self.sheet,
            "headers": self.headers,
            "data_rows": self.data_rows,
            "col_types": {self.headers[i]: t for i, t in enumerate(self.col_types.values())},
            "col_stats": {self.headers[i]: s for i, s in enumerate(self.col_stats.values())},
        }
    
    def close(self):
        if self.wb:
            self.wb.close()


class ChartRecommender:
    """图表推荐器"""
    
    # 数据特征 → 推荐图表
    RECOMMENDATION_RULES = [
        {
            "name": "时间序列折线图",
            "condition": lambda p: any(t == "datetime" for t in p["col_types"].values()) and 
                         any(t == "numeric" for t in p["col_types"].values()),
            "chart_type": "line",
            "description": "展示数据随时间的变化趋势",
            "priority": 100,
        },
        {
            "name": "类别对比柱状图",
            "condition": lambda p: sum(1 for t in p["col_types"].values() if t == "category") >= 1 and
                         sum(1 for t in p["col_types"].values() if t == "numeric") >= 1,
            "chart_type": "bar",
            "description": "比较不同类别之间的数值差异",
            "priority": 90,
        },
        {
            "name": "占比饼图",
            "condition": lambda p: sum(1 for t in p["col_types"].values() if t == "category") >= 1 and
                         sum(1 for t in p["col_types"].values() if t == "numeric") == 1 and
                         p["data_rows"] <= 10,
            "chart_type": "pie",
            "description": "展示各部分占整体的比例（建议 ≤10 个类别）",
            "priority": 80,
        },
        {
            "name": "相关性散点图",
            "condition": lambda p: sum(1 for t in p["col_types"].values() if t == "numeric") >= 2,
            "chart_type": "scatter",
            "description": "分析两个数值变量之间的相关性",
            "priority": 70,
        },
        {
            "name": "堆叠面积图",
            "condition": lambda p: any(t == "datetime" for t in p["col_types"].values()) and
                         sum(1 for t in p["col_types"].values() if t == "numeric") >= 2,
            "chart_type": "area",
            "description": "展示多个时间序列的累积变化",
            "priority": 85,
        },
        {
            "name": "多变量热力图",
            "condition": lambda p: sum(1 for t in p["col_types"].values() if t == "numeric") >= 3 and
                         p["data_rows"] >= 5,
            "chart_type": "heatmap",
            "description": "展示多个变量之间的相关强度",
            "priority": 60,
        },
    ]
    
    def __init__(self):
        self.hw = get_hardware_info()
    
    def recommend(self, profile: Dict) -> List[Dict]:
        """根据数据画像推荐图表"""
        recommendations = []
        
        for rule in self.RECOMMENDATION_RULES:
            try:
                if rule["condition"](profile):
                    recommendations.append({
                        "name": rule["name"],
                        "chart_type": rule["chart_type"],
                        "description": rule["description"],
                        "priority": rule["priority"],
                    })
            except Exception:
                pass
        
        # 按优先级排序
        recommendations.sort(key=lambda x: x["priority"], reverse=True)
        return recommendations
    
    def auto_generate(self, filepath: str, sheet: str = "Sheet1",
                      output_path: str = "", chart_type: str = "") -> Dict:
        """自动生成图表"""
        if not HAS_OPENPYXL:
            return {"success": False, "error": "openpyxl 未安装，请运行: pip install openpyxl"}
        
        # 分析数据
        analyzer = DataAnalyzer(filepath, sheet)
        if not analyzer.load():
            return {"success": False, "error": "无法加载 Excel 文件"}
        
        profile = analyzer.get_profile()
        analyzer.close()
        
        # 推荐图表
        if not chart_type:
            recommendations = self.recommend(profile)
            if not recommendations:
                return {"success": False, "error": "无法识别数据特征，无法推荐图表"}
            chart_type = recommendations[0]["chart_type"]
        
        # 生成图表
        return self._generate_chart(filepath, sheet, chart_type, output_path or filepath, profile)
    
    def _generate_chart(self, filepath: str, sheet: str, chart_type: str,
                        output_path: str, profile: Dict) -> Dict:
        """生成具体图表"""
        try:
            wb = openpyxl.load_workbook(filepath)
            if sheet not in wb.sheetnames:
                sheet = wb.sheetnames[0]
            ws = wb[sheet]
            
            headers = profile["headers"]
            rows = profile["data_rows"]
            
            # 确定数据列
            cat_col = None  # 类别/X轴
            val_cols = []  # 数值/Y轴
            
            for i, (h, t) in enumerate(zip(headers, profile["col_types"].values())):
                if t in ("category", "datetime") and cat_col is None:
                    cat_col = i
                elif t == "numeric":
                    val_cols.append(i)
            
            if cat_col is None:
                cat_col = 0  # 默认第一列
            if not val_cols:
                val_cols = [1] if len(headers) > 1 else [0]
            
            # 创建图表
            chart = self._create_chart(chart_type)
            if not chart:
                return {"success": False, "error": f"不支持的图表类型: {chart_type}"}
            
            # 设置数据范围
            data_end_row = rows + 1  # +1 for header
            
            # 类别轴
            cats = Reference(ws, min_col=cat_col + 1, min_row=2, max_row=data_end_row)
            
            for i, val_idx in enumerate(val_cols[:3]):  # 最多 3 个序列
                data = Reference(ws, min_col=val_idx + 1, min_row=1, max_row=data_end_row)
                chart.add_data(data, titles_from_data=True)
            
            chart.set_categories(cats)
            chart.title = f"{Path(filepath).stem} - {chart_type}图表"
            chart.style = 10
            
            # 添加到工作表
            ws.add_chart(chart, f"{chr(65 + len(val_cols) + 2)}2")
            
            wb.save(output_path)
            wb.close()
            
            return {
                "success": True,
                "chart_type": chart_type,
                "output": output_path,
                "categories_col": headers[cat_col] if cat_col < len(headers) else "未知",
                "value_cols": [headers[v] for v in val_cols[:3] if v < len(headers)],
            }
        except Exception as e:
            return {"success": False, "error": f"图表生成失败: {str(e)}"}
    
    def _create_chart(self, chart_type: str):
        """创建图表对象"""
        chart_map = {
            "line": LineChart,
            "bar": BarChart,
            "pie": PieChart,
            "scatter": ScatterChart,
            "area": AreaChart,
            "heatmap": BarChart,  # openpyxl 无热力图，降级为柱状图
        }
        cls = chart_map.get(chart_type)
        return cls() if cls else None


def _cli():
    """CLI 入口"""
    import argparse
    
    parser = argparse.ArgumentParser(description="数据图表生成器 v4.6.0")
    sub = parser.add_subparsers(dest="command", required=True)
    
    # analyze
    p = sub.add_parser("analyze", help="分析数据特征")
    p.add_argument("--file", required=True)
    p.add_argument("--sheet", default="Sheet1")
    
    # recommend
    p = sub.add_parser("recommend", help="推荐图表类型")
    p.add_argument("--file", required=True)
    p.add_argument("--sheet", default="Sheet1")
    
    # generate
    p = sub.add_parser("generate", help="生成图表")
    p.add_argument("--file", required=True)
    p.add_argument("--sheet", default="Sheet1")
    p.add_argument("--type", default="", help="图表类型（不指定则自动推荐）")
    p.add_argument("--output", default="", help="输出路径")
    
    # auto
    p = sub.add_parser("auto", help="一键分析+推荐+生成")
    p.add_argument("--file", required=True)
    p.add_argument("--sheet", default="Sheet1")
    p.add_argument("--output", default="")
    p.add_argument("--type", default="")
    
    args = parser.parse_args()
    
    if args.command == "analyze":
        a = DataAnalyzer(args.file, args.sheet)
        if a.load():
            print(json.dumps(a.get_profile(), ensure_ascii=False, default=str))
            a.close()
        else:
            print(json.dumps({"success": False, "error": "加载失败"}, ensure_ascii=False))
    
    elif args.command == "recommend":
        a = DataAnalyzer(args.file, args.sheet)
        if a.load():
            profile = a.get_profile()
            r = ChartRecommender()
            recs = r.recommend(profile)
            print(json.dumps({"success": True, "recommendations": recs}, ensure_ascii=False, default=str))
            a.close()
        else:
            print(json.dumps({"success": False, "error": "加载失败"}, ensure_ascii=False))
    
    elif args.command == "generate":
        r = ChartRecommender()
        result = r.auto_generate(args.file, args.sheet, args.output, args.type)
        print(json.dumps(result, ensure_ascii=False, default=str))
    
    elif args.command == "auto":
        r = ChartRecommender()
        result = r.auto_generate(args.file, args.sheet, args.output, args.type)
        print(json.dumps(result, ensure_ascii=False, default=str))


if __name__ == "__main__":
    _cli()
